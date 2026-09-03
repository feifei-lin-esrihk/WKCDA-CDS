r"""
Create file geodatabases (GDBs) from the proposed data structure workbook.
 
Workflow:
    1. Read the Excel workbook (one sheet per plan, e.g. P3 / P4 / P7-18).
    2. Parse each sheet into feature classes with their field specs
       (name, alias, type, length) and default values, taken both from the
       per-field "Default" column and from the FC-level columns
       (Plan / Source Filename / Feature Class / Feature Type).
    3. Expand plan ranges: a sheet like "P7-18" (Plan column = P7-18)
       produces one GDB per plan in the range, i.e. P7.gdb ... P18.gdb,
       each containing the sheet's feature classes with SOURCE_PLAN
       defaulted to that specific plan. SHEETS may also list single plan
       names (e.g. "P7") to build only that plan's GDB from the range
       sheet, so teammates can each generate their own subset.
    4. Create each GDB, feature class and field with the specified alias,
       type and length; apply field defaults; and attach an auto-increment
       attribute rule to FEAT_ID (database sequence + Insert rule).
 
Run with the ArcGIS Pro Python environment, e.g.:
    "C:\Program Files\ArcGIS\Pro\bin\Python\envs\arcgispro-py3\python.exe" create_gdbs.py
"""
 
import os
import re
import sys
import arcpy
import openpyxl
 
# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
 
# Full path to the data structure workbook (the source of all field specs).
XLSX_PATH = r"C:\Users\Sam\Git\WKCDA\WKCDA_Proposed Data Structure.xlsx"
 
# Folder where the GDBs will be created.
OUTPUT_FOLDER = r"C:\Users\Sam\Git\WKCDA"
 
# Workbook entries to process. Each entry may be either a sheet name or a
# single plan name:
#   - A sheet name (e.g. "P3", "P4") builds one GDB per plan the sheet
#     defines; a range Plan value like "P7-18" expands into P7.gdb..P18.gdb.
#   - A single plan name (e.g. "P7") finds the sheet that covers it (e.g.
#     "P7-18") and builds ONLY that plan's GDB, so teammates can each run
#     their own subset, e.g. SHEETS = ["P7", "P8"].
SHEETS = ["P7", "P18"]
 
# Spatial reference for all new feature classes, by well-known ID (WKID).
# 2326 = Hong Kong 1980 Grid Coordinate System; change if another CRS is used.
SPATIAL_REFERENCE_WKID = 2326
 
# When True, FEAT_ID is auto-incremented on every new record using a
# database sequence + attribute rule (requires ArcGIS Pro 2.9 or later;
# the feature class also gets GlobalIDs, which attribute rules depend on).
AUTO_INCREMENT_FEAT_ID = True
 
# Fields listed here are created as NON_NULLABLE. FEAT_ID is safe to make
# non-nullable because the Insert attribute rule always populates it during
# editing; batch loads via InsertCursor must then supply a value explicitly.
NON_NULLABLE_FIELDS = {"FEAT_ID"}
 
# Database sequence settings for FEAT_ID (per FC), following Esri KB
# "Configure a unique ID that auto-increments by a certain number":
#   FEAT_ID_SEQUENCE_START     -> first value returned by the sequence
#   FEAT_ID_SEQUENCE_INCREMENT -> step between consecutive values
# Example: start 1000, increment 10 -> 1000, 1010, 1020, ...
FEAT_ID_SEQUENCE_START = 1
FEAT_ID_SEQUENCE_INCREMENT = 1
 
# Per the same Esri KB, Is Editable should be unchecked so the database
# values cannot be overridden by editors / field workers. Set True if you
# need to type a specific FEAT_ID manually on occasion.
FEAT_ID_RULE_EDITABLE = False
 
# Translate spreadsheet field types (as written in the "Field Type" column)
# into arcpy field type keywords. Unknown types fall back to TEXT.
FIELD_TYPE_MAP = {
    "Long": "LONG",
    "Short": "SHORT",
    "Text": "TEXT",
    "Float": "FLOAT",
    "Double": "DOUBLE",
    "Date": "DATE",
}
 
# Translate spreadsheet geometry types (the "Feature Type" column) into
# arcpy geometry keywords; defaults to POLYGON when not recognised.
GEOMETRY_TYPE_MAP = {
    "POLYGON": "POLYGON",
    "POLYLINE": "POLYLINE",
    "LINE": "POLYLINE",
    "POINT": "POINT",
    "MULTIPOINT": "MULTIPOINT",
}
 
# Placeholder used in the workbook for "no value"; such cells are treated
# as empty when deriving FC-level defaults (e.g. Source Filename "-").
EMPTY_PLACEHOLDERS = {"-", "n/a", "na", "nil", "none"}
 
 
# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
 
def clean(value):
    """Normalise a cell value: None -> "", whole-number floats -> int.
 
    openpyxl returns numbers as floats (e.g. 50.0 for a field length),
    so integer-valued floats are converted to plain ints for cleaner output.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value
 
 
def header_index(header_values, name, last=False):
    """Find a column index by its header text (case-insensitive).
 
    Returns the first matching index, or the last one when last=True.
    "Default" is looked up with last=True because some sheets (e.g. P2)
    carry a duplicated Default header and the values sit in the rightmost
    copy. Returns None when the sheet has no such column (e.g. Common).
    """
    matches = [i for i, v in enumerate(header_values)
               if str(v).strip().lower() == name.strip().lower()]
    if not matches:
        return None
    return matches[-1] if last else matches[0]
 
 
def is_empty(value):
    """True when a cell value is blank or a placeholder like '-'."""
    return value == "" or str(value).strip().lower() in EMPTY_PLACEHOLDERS
 
 
def sanitize_name(name, max_length=160):
    """Make a workbook name safe as a geodatabase table/GDB name.
 
    Runs of characters outside [A-Za-z0-9_] (spaces, punctuation, Chinese
    characters, ...) become a single underscore, e.g. "MRCP_LABEL PT" ->
    "MRCP_LABEL_PT". Names starting with a digit are prefixed with "FC_"
    and over-long names are trimmed to max_length.
    """
    sanitized = re.sub(r"[^A-Za-z0-9_]+", "_", str(name).strip())
    sanitized = re.sub(r"_+", "_", sanitized).strip("_")
    if not sanitized:
        sanitized = "UNNAMED"
    if sanitized[0].isdigit():
        sanitized = "FC_" + sanitized
    return sanitized[:max_length]
 
 
def looks_like_fc_name(text):
    """Heuristic separating real feature class names from workbook remarks.
 
    Real names are single tokens of letters/digits/underscores, or short
    all-caps multi-token names such as "MRCP_LABEL PT" (a typo for
    MRCP_LABEL_PT). Prose remarks like "Data quite messy." or "Some points
    out of HK." contain lowercase words / punctuation and are rejected so
    they are never turned into feature classes.
    """
    if re.match(r"^[A-Za-z][A-Za-z0-9_]*$", text):
        return True
    if len(text) <= 40 and re.match(r"^[A-Z0-9][A-Za-z0-9_]*( [A-Z0-9][A-Za-z0-9_]*)*$", text):
        return True
    return False
 
 
def expand_plans(plan_value):
    """Expand a Plan column value into the list of GDB names it represents.
 
    A range value like "P7-18" expands to ["P7", "P8", ..., "P18"]; any
    other value (e.g. "P3") is returned as a single-item list.
    """
    match = re.match(r"^([A-Za-z]+)(\d+)\s*-\s*(\d+)$", str(plan_value).strip())
    if match:
        prefix, start, end = match.group(1), int(match.group(2)), int(match.group(3))
        if start < end:
            return ["{0}{1}".format(prefix, i) for i in range(start, end + 1)]
    return [str(plan_value).strip()]
 
 
def default_value_for_type(value, field_type):
    """Convert a workbook default value to the type arcpy expects.
 
    Numeric fields (LONG/SHORT/DOUBLE/FLOAT) get int/float values; TEXT
    fields always get a string (e.g. numeric cell 11260 -> "11260").
    Unconvertible values are passed through unchanged so AssignDefaultToField
    raises a clear error instead of silently guessing.
    """
    if field_type in ("LONG", "SHORT"):
        try:
            return int(value)
        except (TypeError, ValueError):
            return value
    if field_type in ("DOUBLE", "FLOAT"):
        try:
            return float(value)
        except (TypeError, ValueError):
            return value
    return str(value)
 
 
# ---------------------------------------------------------------------------
# Workbook parsing
# ---------------------------------------------------------------------------
 
def parse_sheet(ws):
    """Parse one workbook sheet into a list of feature class definitions.
 
    Column positions are located from the header row (row 2) by name, so
    extra or duplicated columns are tolerated:
        Plan             -> GDB names via expand_plans; default for SOURCE_PLAN
        Source Filename  -> default value for SOURCE_DOC
        Feature Class    -> starts a new FC block; default value for FC_NAME
        Feature Type     -> geometry type; default for SOURCE_FEAT_TYPE
        Field / Field Alias / Field Type / Field Length -> the field spec
        Default          -> explicit per-field default value
 
    A marker row such as "(add to all fearues)" (field name starting with
    "(") marks that the NEXT field listed belongs to every feature class in
    the sheet (e.g. COMP_YEAR). Such fields are collected separately and
    appended to all feature classes at the end, skipping duplicates. Note
    that COMP_YEAR may also appear inline inside each FC block; the dedupe
    step keeps only the first occurrence.
 
    Returns a list of dicts:
        [{"name": ..., "geometry": ..., "plan": ...,
          "defaults": {field: value, ...},
          "fields": [{name, alias, type, length, default}]}]
    """
    # Resolve the column layout from the header row (row 2).
    header = [clean(c.value) for c in ws[2]]
    cols = {
        "plan": header_index(header, "Plan"),
        "src": header_index(header, "Source Filename"),
        "fc": header_index(header, "Feature Class"),
        "ftype": header_index(header, "Feature Type"),
        "field": header_index(header, "Field"),
        "alias": header_index(header, "Field Alias"),
        "type": header_index(header, "Field Type"),
        "length": header_index(header, "Field Length"),
        "default": header_index(header, "Default", last=True),
    }
 
    def cell(values, idx):
        return values[idx] if idx is not None and idx < len(values) else ""
 
    feature_classes = []
    current_fc = None
    add_to_all = False
    global_fields = []
    fc_by_name = {}
 
    for row in ws.iter_rows(min_row=3):
        values = [clean(c.value) for c in row]
        plan = str(cell(values, cols["plan"])).strip()
        src_file = str(cell(values, cols["src"])).strip()
        fc_name = str(cell(values, cols["fc"])).strip()
        fc_type = str(cell(values, cols["ftype"])).strip()
        field_name = str(cell(values, cols["field"])).strip()
        field_alias = str(cell(values, cols["alias"])).strip()
        field_type = str(cell(values, cols["type"])).strip()
        field_length = cell(values, cols["length"])
        field_default = cell(values, cols["default"])
 
        # Skip blank rows / rows without a field definition.
        if field_name == "":
            continue
 
        # Marker row, e.g. "(add to all fearues)": the next field is global.
        if field_name.startswith("("):
            add_to_all = True
            continue
 
        # Build the field specification for this row. An explicit Default
        # column value is carried on the field; it is applied in build_gdb.
        field_spec = {
            "name": field_name,
            "alias": field_alias or field_name,  # default alias = field name
            "type": FIELD_TYPE_MAP.get(field_type.upper(), FIELD_TYPE_MAP.get(field_type, "TEXT")),
            "length": int(field_length) if isinstance(field_length, int) or str(field_length).isdigit() else None,
            "default": None if is_empty(field_default) else field_default,
        }
 
        # A non-empty Feature Class cell starts a new feature class block;
        # subsequent rows belong to it until the next block starts. Prose
        # remarks in that column are ignored (no block is started, any
        # fields on such a row stay with the previous FC), real names are
        # sanitized, and duplicate blocks for the same FC are merged.
        if fc_name and looks_like_fc_name(fc_name):
            fc_name = sanitize_name(fc_name)
            if fc_name in fc_by_name:
                current_fc = fc_by_name[fc_name]
                print("INFO: duplicate feature class block {0} in sheet, "
                      "fields merged".format(fc_name))
            else:
                # FC-level column values are captured as field defaults
                # (placeholders like "-" are treated as "no default").
                defaults = {"FC_NAME": fc_name}
                if not is_empty(plan):
                    defaults["SOURCE_PLAN"] = plan
                if not is_empty(src_file):
                    defaults["SOURCE_DOC"] = src_file
                if fc_type:
                    defaults["SOURCE_FEAT_TYPE"] = fc_type
                current_fc = {
                    "name": fc_name,
                    "geometry": GEOMETRY_TYPE_MAP.get(fc_type.upper(), "POLYGON"),
                    "plan": plan or None,
                    "defaults": defaults,
                    "fields": [],
                }
                fc_by_name[fc_name] = current_fc
                feature_classes.append(current_fc)
        elif fc_name:
            print("INFO: remark in Feature Class column ignored: {0!r}".format(fc_name))
 
        # Route the field: global fields go to a shared list, everything
        # else is appended to the current feature class. Duplicate field
        # names inside one FC (e.g. from merged duplicate blocks) are
        # skipped, keeping the first definition.
        if add_to_all or current_fc is None:
            global_fields.append(field_spec)
            add_to_all = False
        elif not any(f["name"] == field_spec["name"] for f in current_fc["fields"]):
            current_fc["fields"].append(field_spec)
 
    # Append each global field (e.g. COMP_YEAR) to every feature class,
    # unless a field with the same name is already defined there.
    for fc in feature_classes:
        existing = {f["name"] for f in fc["fields"]}
        for g in global_fields:
            if g["name"] not in existing:
                fc["fields"].append(g)
 
    return feature_classes
 
 
# ---------------------------------------------------------------------------
# GDB creation
# ---------------------------------------------------------------------------
 
def add_feat_id_rule(gdb_path, fc_path, fc_name):
    """Make FEAT_ID auto-increment whenever a new record is inserted.
 
    Three steps are required for a file GDB:
        1. AddGlobalIDs        -> attribute rules only work on datasets that
                                  have a GlobalID field.
        2. CreateDatabaseSequence -> a counter stored in the GDB, one per
                                  feature class, using the configured start
                                  and increment values.
        3. AddAttributeRule    -> an Arcade rule bound to FEAT_ID that fires
                                  on Insert and assigns the next sequence
                                  value, so every new feature (digitised in
                                  Pro or added via a feature service) gets
                                  FEAT_ID auto-numbered. The rule is created
                                  with "Exclude from application evaluation"
                                  enabled, which is mandatory for sequence
                                  rules in ArcGIS Pro 3.5+ (ERROR 002920);
                                  the value is assigned when edits are saved
                                  rather than at feature creation.
 
    The rule is editable per FEAT_ID_RULE_EDITABLE (Esri KB recommends
    leaving it non-editable so field workers cannot override the IDs).
 
    The AddAttributeRule call is built from arcpy.GetParameterInfo at runtime:
    parameter names differ between ArcGIS Pro versions (e.g. some versions
    require a rule "Type" parameter), so each desired value is matched to the
    actual tool parameter by internal name or display name. If the version
    has a required Type parameter, the "calculation" option is chosen
    automatically, since this rule assigns a value to a field.
    """
    arcpy.management.AddGlobalIDs(fc_path)
 
    seq_name = "{0}_FEATID_SEQ".format(fc_name[:20])
    arcpy.management.CreateDatabaseSequence(
        gdb_path, seq_name, FEAT_ID_SEQUENCE_START, FEAT_ID_SEQUENCE_INCREMENT)
 
    rule_name = "RULE_{0}_FEAT_ID".format(fc_name[:20])
    expression = "return NextSequenceValue('{0}')".format(seq_name)
 
    # Desired argument values, keyed by every plausible parameter name.
    desired = {
        "name": rule_name,
        "rule_name": rule_name,
        "script_expression": expression,
        "arcade_expression": expression,
        "is_editable": FEAT_ID_RULE_EDITABLE,
        "triggering_events": "Insert",
        "field": "FEAT_ID",
        # Required for sequence-based rules (ERROR 002920, ArcGIS Pro 3.5):
        # NextSequenceValue is evaluated on save / server-side, not by the
        # client during the edit operation.
        "exclude_from_client_evaluation": True,
        "description": "Auto-populate FEAT_ID with the next sequence value",
    }
 
    # Map the desired values onto this Pro version's actual parameters.
    kwargs = {}
    for p in arcpy.GetParameterInfo("management.AddAttributeRule"):
        for dname, dvalue in desired.items():
            if p.name.lower() == dname.lower() or (p.displayName or "").lower() == dname.lower():
                kwargs[p.name] = dvalue
                break
        else:
            # Required rule-type parameter (not all versions have it):
            # pick the "calculation" choice, or the first available choice.
            if p.parameterType == "Required" and p.name.lower() in ("type", "rule_type"):
                choices = list(getattr(p.filter, "list", None) or [])
                calc = [c for c in choices if "calc" in str(c).lower()]
                kwargs[p.name] = calc[0] if calc else (choices[0] if choices else "calculation")
 
    arcpy.management.AddAttributeRule(fc_path, **kwargs)
 
 
def build_gdb(gdb_name, feature_classes):
    """Create <gdb_name>.gdb and populate it with the parsed feature classes.
 
    An existing GDB of the same name is deleted first so the script can be
    re-run safely (idempotent). For each feature class:
        1. CreateFeatureclass -> empty FC with the configured spatial reference
           (OBJECTID and Shape are added automatically by ArcGIS).
        2. AddField for every parsed field, applying its alias; TEXT fields
           also get their specified length (default 255 when missing), and
           fields in NON_NULLABLE_FIELDS are created NON_NULLABLE.
        3. AssignDefaultToField for the effective defaults: FC-level workbook
           values (FC_NAME / SOURCE_PLAN / SOURCE_DOC / SOURCE_FEAT_TYPE)
           overridden by any explicit "Default" column values. Values are
           converted to the field's type, and TEXT defaults that would
           exceed the field length are skipped with a warning.
        4. If AUTO_INCREMENT_FEAT_ID is on and the FC has a FEAT_ID field,
           attach an auto-increment rule to it (see add_feat_id_rule).
    Finally a summary of every FC and its fields is printed for checking.
    """
    gdb_name = sanitize_name(gdb_name)
    gdb_path = os.path.join(OUTPUT_FOLDER, gdb_name + ".gdb")
 
    # Recreate the GDB from scratch so re-runs always match the workbook.
    if arcpy.Exists(gdb_path):
        arcpy.management.Delete(gdb_path)
    arcpy.management.CreateFileGDB(OUTPUT_FOLDER, gdb_name + ".gdb")
 
    spatial_ref = arcpy.SpatialReference(SPATIAL_REFERENCE_WKID)
 
    for fc in feature_classes:
        fc_path = os.path.join(gdb_path, fc["name"])
 
        # Empty feature class in the GDB; geometry only, no custom fields yet.
        arcpy.management.CreateFeatureclass(
            gdb_path,
            fc["name"],
            fc["geometry"],
            spatial_reference=spatial_ref,
        )
 
        # Add each field from the workbook spec. field_length is only valid
        # for TEXT fields, so it is only passed for those. Fields listed in
        # NON_NULLABLE_FIELDS are created with NULLs disallowed.
        for f in fc["fields"]:
            kwargs = {"field_alias": f["alias"]}
            if f["type"] == "TEXT":
                kwargs["field_length"] = f["length"] if f["length"] else 255
            if f["name"] in NON_NULLABLE_FIELDS:
                kwargs["field_is_nullable"] = "NON_NULLABLE"
            arcpy.management.AddField(fc_path, f["name"], f["type"], **kwargs)
 
        # Effective defaults = FC-level values overridden by explicit
        # per-field "Default" column values.
        effective = dict(fc.get("defaults", {}))
        for f in fc["fields"]:
            if f.get("default") is not None:
                effective[f["name"]] = f["default"]
 
        # Apply defaults for fields that exist on this FC, converting each
        # value to the field's type; TEXT values exceeding the field length
        # are skipped with a warning instead of being truncated.
        applied_defaults = []
        for field_name, value in effective.items():
            spec = next((f for f in fc["fields"] if f["name"] == field_name), None)
            if spec is None:
                continue
            typed = default_value_for_type(value, spec["type"])
            if spec["type"] == "TEXT" and len(str(typed)) > (spec["length"] or 255):
                print("    WARNING: default for {0} ({1} chars) exceeds field "
                      "length {2}, skipped".format(field_name, len(str(typed)),
                                                   spec["length"] or 255))
                continue
            arcpy.management.AssignDefaultToField(fc_path, field_name, typed)
            applied_defaults.append("{0}={1!r}".format(field_name, typed))
        if applied_defaults:
            print("    Defaults: {0}".format(", ".join(applied_defaults)))
 
        # Auto-increment FEAT_ID via sequence + attribute rule, when enabled.
        # Failures (e.g. old ArcGIS Pro) are reported but do not stop the run;
        # the warning includes the tool's parameter signature to help diagnose
        # version differences.
        if AUTO_INCREMENT_FEAT_ID and any(f["name"] == "FEAT_ID" for f in fc["fields"]):
            try:
                add_feat_id_rule(gdb_path, fc_path, fc["name"])
                print("    FEAT_ID auto-increment enabled (sequence + Insert rule)")
            except Exception as ex:
                try:
                    sig = "; ".join("{0}={1}".format(p.name, p.parameterType)
                                    for p in arcpy.GetParameterInfo("management.AddAttributeRule"))
                    detail = " | AddAttributeRule params: {0}".format(sig)
                except Exception:
                    detail = ""
                print("    WARNING: could not enable FEAT_ID auto-increment: {0}{1}".format(ex, detail))
 
        # Print a summary of what was created for visual verification.
        print("Created {0}\\{1} ({2}) with {3} fields".format(
            gdb_path, fc["name"], fc["geometry"], len(fc["fields"])))
        for f in fc["fields"]:
            length = "({0})".format(f["length"]) if f["type"] == "TEXT" else ""
            dflt = " [default: {0!r}]".format(f["default"]) if f.get("default") is not None else ""
            print("    {0:<20} {1:<8}{2:<8} {3}{4}".format(
                f["name"], f["type"], length, f["alias"], dflt))
 
 
# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------
 
def main():
    if not os.path.exists(XLSX_PATH):
        sys.exit("Excel file not found: {0}".format(XLSX_PATH))
    if not os.path.exists(OUTPUT_FOLDER):
        os.makedirs(OUTPUT_FOLDER)
 
    # data_only=True reads cached cell values instead of formulas.
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
 
    # Cache parse results so range sheets are only parsed once.
    parsed_cache = {}
 
    def parsed(sheet):
        if sheet not in parsed_cache:
            parsed_cache[sheet] = parse_sheet(wb[sheet])
        return parsed_cache[sheet]
 
    def covers(sheet, plan_name):
        """True when the sheet produces the given plan name (directly or
        through a range Plan value such as P7-18)."""
        for fc in parsed(sheet):
            if plan_name in expand_plans(fc.get("plan") or sheet):
                return True
        return False
 
    def build_plan(gdb_name, feature_classes, built):
        """Build one GDB with SOURCE_PLAN defaulted to the specific plan."""
        expanded_fcs = []
        for fc in feature_classes:
            clone = dict(fc)
            clone["defaults"] = dict(fc["defaults"])
            clone["defaults"]["SOURCE_PLAN"] = gdb_name
            expanded_fcs.append(clone)
        build_gdb(gdb_name, expanded_fcs)
        built.add(gdb_name)
 
    built = set()
    for requested in SHEETS:
        if requested in wb.sheetnames:
            # Sheet entry: build every GDB the sheet defines.
            feature_classes = parsed(requested)
            if not feature_classes:
                print("No feature classes found in sheet: {0}".format(requested))
                continue
 
            # Group the FCs by their raw Plan value; a sheet may in theory
            # mix several plans. FCs without a plan fall back to the sheet
            # name, which expand_plans() keeps as a single name.
            groups = {}
            for fc in feature_classes:
                groups.setdefault(fc.get("plan") or requested, []).append(fc)
 
            for plan_value, fcs in groups.items():
                gdb_names = expand_plans(plan_value)
                if len(gdb_names) > 1:
                    print("Expanding plan range {0} -> {1}..{2} ({3} GDBs)".format(
                        plan_value, gdb_names[0], gdb_names[-1], len(gdb_names)))
 
                for gdb_name in gdb_names:
                    if gdb_name in built:
                        print("GDB {0} already built this run, skipped".format(gdb_name))
                        continue
                    build_plan(gdb_name, fcs, built)
        else:
            # Plan-name entry: find the sheet that covers it and build only
            # that plan's GDB (e.g. "P7" -> the P7-18 sheet, P7.gdb only).
            owner = next((s for s in wb.sheetnames if covers(s, requested)), None)
            if owner is None:
                print("No sheet or plan found for '{0}', skipped".format(requested))
                continue
            if requested in built:
                print("GDB {0} already built this run, skipped".format(requested))
                continue
            fcs = [fc for fc in parsed(owner)
                   if requested in expand_plans(fc.get("plan") or owner)]
            build_plan(requested, fcs, built)
    print("Done.")
 
 
if __name__ == "__main__":
    main()
